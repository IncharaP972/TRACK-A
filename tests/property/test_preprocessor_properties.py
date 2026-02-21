"""
Property-based tests for the Excel preprocessor.

Feature: latspace-excel-parser
"""

import pytest
from hypothesis import given, strategies as st, assume
from openpyxl import Workbook
from app.core.preprocessor import find_header_row
from app.schema.models import TableStructure


# Property 10: Header Row Detection Consistency
# **Validates: Requirements 7.4**


@given(
    # Generate a header row index between 1 and 10
    header_row_idx=st.integers(min_value=1, max_value=10),
    # Generate number of columns
    num_columns=st.integers(min_value=4, max_value=20),
    # Generate number of data rows after header
    num_data_rows=st.integers(min_value=0, max_value=5),
)
def test_header_row_detection_consistency(header_row_idx, num_columns, num_data_rows):
    """
    Property 10: Header Row Detection Consistency
    For any detected header row at index N, the table structure should have
    data_start_row set to N+1.
    
    **Validates: Requirements 7.4**
    """
    # Create a workbook with a clear header row
    wb = Workbook()
    sheet = wb.active
    
    # Add rows before the header (if any) with numeric data
    for row_idx in range(1, header_row_idx):
        for col_idx in range(1, num_columns + 1):
            # Add numeric values or empty cells (not strings)
            if col_idx % 2 == 0:
                sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
            # Leave other cells empty
    
    # Add the header row with >3 string values to trigger heuristic
    header_values = [f"Header_{i}" for i in range(1, num_columns + 1)]
    for col_idx, header_val in enumerate(header_values, start=1):
        sheet.cell(row=header_row_idx, column=col_idx, value=header_val)
    
    # Add data rows after the header
    for data_row_offset in range(1, num_data_rows + 1):
        data_row_idx = header_row_idx + data_row_offset
        for col_idx in range(1, num_columns + 1):
            sheet.cell(row=data_row_idx, column=col_idx, value=data_row_offset * col_idx)
    
    # Detect header row (without LLM service)
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 10: data_start_row = header_row_index + 1
    assert isinstance(table_structure, TableStructure), \
        "find_header_row should return a TableStructure"
    
    assert table_structure.header_row_index == header_row_idx, \
        f"Expected header_row_index={header_row_idx}, got {table_structure.header_row_index}"
    
    assert table_structure.data_start_row == header_row_idx + 1, \
        f"Expected data_start_row={header_row_idx + 1}, got {table_structure.data_start_row}"
    
    # Additional verification: data_start_row should always be header_row_index + 1
    assert table_structure.data_start_row == table_structure.header_row_index + 1, \
        "data_start_row must equal header_row_index + 1"


@given(
    # Generate multiple candidate rows
    num_candidates=st.integers(min_value=2, max_value=5),
    num_columns=st.integers(min_value=4, max_value=15),
)
def test_header_row_detection_consistency_multiple_candidates(num_candidates, num_columns):
    """
    Property 10: Header Row Detection Consistency (Multiple candidates variant)
    When multiple rows have >3 strings, the selected header row should still
    satisfy data_start_row = header_row_index + 1.
    
    **Validates: Requirements 7.4**
    """
    # Create a workbook with multiple candidate rows
    wb = Workbook()
    sheet = wb.active
    
    # Add multiple rows with string values (all candidates)
    # The one with the most strings should be selected
    max_strings = 0
    expected_header_idx = 1
    
    for row_idx in range(1, num_candidates + 1):
        # Each row has a different number of strings
        num_strings = 4 + row_idx  # Ensure >3 strings for all
        
        if num_strings > max_strings:
            max_strings = num_strings
            expected_header_idx = row_idx
        
        for col_idx in range(1, min(num_strings, num_columns) + 1):
            sheet.cell(row=row_idx, column=col_idx, value=f"String_{row_idx}_{col_idx}")
        
        # Fill remaining columns with numbers
        for col_idx in range(num_strings + 1, num_columns + 1):
            sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
    
    # Detect header row
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 10
    assert table_structure.data_start_row == table_structure.header_row_index + 1, \
        f"data_start_row ({table_structure.data_start_row}) must equal " \
        f"header_row_index ({table_structure.header_row_index}) + 1"


@given(
    header_row_idx=st.integers(min_value=1, max_value=8),
    num_columns=st.integers(min_value=5, max_value=12),
)
def test_header_row_detection_consistency_with_empty_rows(header_row_idx, num_columns):
    """
    Property 10: Header Row Detection Consistency (Empty rows variant)
    Even with empty rows before the header, the property should hold.
    
    **Validates: Requirements 7.4**
    """
    # Create a workbook with empty rows before header
    wb = Workbook()
    sheet = wb.active
    
    # Leave rows before header_row_idx empty
    
    # Add the header row with many string values
    for col_idx in range(1, num_columns + 1):
        sheet.cell(row=header_row_idx, column=col_idx, value=f"Column_{col_idx}")
    
    # Add one data row after header
    for col_idx in range(1, num_columns + 1):
        sheet.cell(row=header_row_idx + 1, column=col_idx, value=col_idx * 10)
    
    # Detect header row
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 10
    assert table_structure.data_start_row == table_structure.header_row_index + 1, \
        f"data_start_row ({table_structure.data_start_row}) must equal " \
        f"header_row_index ({table_structure.header_row_index}) + 1"


# Property 11: Merged Cell Capture
# **Validates: Requirements 7.5**


@given(
    # Generate merged cell ranges
    num_merged_ranges=st.integers(min_value=1, max_value=5),
    base_row=st.integers(min_value=1, max_value=5),
    base_col=st.integers(min_value=1, max_value=5),
)
def test_merged_cell_capture(num_merged_ranges, base_row, base_col):
    """
    Property 11: Merged Cell Capture
    For any Excel file with merged cells, the TableStructure should include
    all merged cell ranges in the merged_cells list.
    
    **Validates: Requirements 7.5**
    """
    # Create a workbook with merged cells
    wb = Workbook()
    sheet = wb.active
    
    # Track expected merged ranges
    expected_merged_ranges = []
    
    # Create merged cell ranges
    for i in range(num_merged_ranges):
        # Calculate merge range (ensure no overlap)
        start_row = base_row + (i * 3)
        end_row = start_row + 1  # Merge 2 rows
        start_col = base_col + (i * 2)
        end_col = start_col + 1  # Merge 2 columns
        
        # Merge the cells
        sheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col
        )
        
        # Add to expected ranges
        expected_merged_ranges.append({
            "min_row": start_row,
            "max_row": end_row,
            "min_col": start_col,
            "max_col": end_col
        })
    
    # Add a clear header row (after all merged cells)
    header_row_idx = base_row + (num_merged_ranges * 3) + 1
    for col_idx in range(1, 10):
        sheet.cell(row=header_row_idx, column=col_idx, value=f"Header_{col_idx}")
    
    # Detect table structure
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 11: All merged cells are captured
    assert isinstance(table_structure.merged_cells, list), \
        "merged_cells should be a list"
    
    assert len(table_structure.merged_cells) == num_merged_ranges, \
        f"Expected {num_merged_ranges} merged cell ranges, got {len(table_structure.merged_cells)}"
    
    # Verify each merged range is captured
    for expected_range in expected_merged_ranges:
        # Check if this range exists in the captured merged cells
        found = any(
            captured["min_row"] == expected_range["min_row"] and
            captured["max_row"] == expected_range["max_row"] and
            captured["min_col"] == expected_range["min_col"] and
            captured["max_col"] == expected_range["max_col"]
            for captured in table_structure.merged_cells
        )
        
        assert found, \
            f"Expected merged range {expected_range} not found in captured merged_cells"


@given(
    # Generate a single large merged cell
    start_row=st.integers(min_value=1, max_value=3),
    start_col=st.integers(min_value=1, max_value=3),
    row_span=st.integers(min_value=2, max_value=5),
    col_span=st.integers(min_value=2, max_value=5),
)
def test_merged_cell_capture_single_large_merge(start_row, start_col, row_span, col_span):
    """
    Property 11: Merged Cell Capture (Single large merge variant)
    For a single large merged cell, it should be captured correctly.
    
    **Validates: Requirements 7.5**
    """
    # Create a workbook with one large merged cell
    wb = Workbook()
    sheet = wb.active
    
    end_row = start_row + row_span - 1
    end_col = start_col + col_span - 1
    
    # Merge the cells
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col
    )
    
    # Add a clear header row after the merged cell
    header_row_idx = end_row + 2
    for col_idx in range(1, 8):
        sheet.cell(row=header_row_idx, column=col_idx, value=f"Col_{col_idx}")
    
    # Detect table structure
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 11
    assert len(table_structure.merged_cells) == 1, \
        f"Expected 1 merged cell range, got {len(table_structure.merged_cells)}"
    
    merged_range = table_structure.merged_cells[0]
    assert merged_range["min_row"] == start_row, \
        f"Expected min_row={start_row}, got {merged_range['min_row']}"
    assert merged_range["max_row"] == end_row, \
        f"Expected max_row={end_row}, got {merged_range['max_row']}"
    assert merged_range["min_col"] == start_col, \
        f"Expected min_col={start_col}, got {merged_range['min_col']}"
    assert merged_range["max_col"] == end_col, \
        f"Expected max_col={end_col}, got {merged_range['max_col']}"


def test_merged_cell_capture_no_merged_cells():
    """
    Property 11: Merged Cell Capture (No merged cells variant)
    For an Excel file with no merged cells, merged_cells list should be empty.
    
    **Validates: Requirements 7.5**
    """
    # Create a workbook with no merged cells
    wb = Workbook()
    sheet = wb.active
    
    # Add a header row
    for col_idx in range(1, 6):
        sheet.cell(row=1, column=col_idx, value=f"Header_{col_idx}")
    
    # Detect table structure
    table_structure = find_header_row(sheet, llm_service=None)
    
    # Verify Property 11: merged_cells should be empty
    assert isinstance(table_structure.merged_cells, list), \
        "merged_cells should be a list"
    assert len(table_structure.merged_cells) == 0, \
        f"Expected 0 merged cells, got {len(table_structure.merged_cells)}"
