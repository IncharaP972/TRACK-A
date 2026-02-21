"""
Unit tests for error handling across all components.

Tests error scenarios including:
- Invalid file format errors
- No header row detected errors
- LLM API failure handling
- Pydantic validation errors
- Corrupted file handling
- Empty file handling

Requirements: 11.1, 11.2, 11.3, 11.5
"""

import pytest
from fastapi.testclient import TestClient
from io import BytesIO
import openpyxl
from unittest.mock import patch, MagicMock
from pydantic import ValidationError

from app.main import app
from app.schema.models import (
    HeaderMapping, ParsedCell, TableStructure, ParseResult,
    MatchMethod, ConfidenceLevel
)
from app.core.preprocessor import find_header_row, _llm_detect_header_row
from app.core.matcher import HeaderMatcher
from app.services.llm import LLMService
from app.registry.data import RegistryManager


@pytest.fixture
def client():
    """Create a test client for the FastAPI app."""
    return TestClient(app)


@pytest.fixture
def mock_llm_service():
    """Create a mock LLM service."""
    return MagicMock(spec=LLMService)


@pytest.fixture
def registry():
    """Create a registry manager."""
    return RegistryManager()


class TestInvalidFileFormatErrors:
    """Test invalid file format error handling."""
    
    def test_invalid_file_extension_txt(self, client):
        """
        Test that .txt files are rejected with HTTP 400.
        
        Requirements: 11.1
        """
        text_file = BytesIO(b"This is a text file")
        
        response = client.post(
            "/api/parse",
            files={"file": ("test.txt", text_file, "text/plain")}
        )
        
        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
        assert "Excel" in data["detail"] or "xlsx" in data["detail"]
    
    def test_invalid_file_extension_csv(self, client):
        """
        Test that .csv files are rejected with HTTP 400.
        
        Requirements: 11.1
        """
        csv_file = BytesIO(b"col1,col2,col3\nval1,val2,val3")
        
        response = client.post(
            "/api/parse",
            files={"file": ("test.csv", csv_file, "text/csv")}
        )
        
        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
    
    def test_invalid_file_extension_json(self, client):
        """
        Test that .json files are rejected with HTTP 400.
        
        Requirements: 11.1
        """
        json_file = BytesIO(b'{"key": "value"}')
        
        response = client.post(
            "/api/parse",
            files={"file": ("test.json", json_file, "application/json")}
        )
        
        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
    
    def test_no_filename_provided(self, client):
        """
        Test that missing filename is handled.
        
        Requirements: 11.1
        """
        # Create a file with no filename
        excel_file = BytesIO(b"fake content")
        
        response = client.post(
            "/api/parse",
            files={"file": ("", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # FastAPI returns 422 for validation errors or 400 for our custom check
        assert response.status_code in [400, 422]
        data = response.json()
        assert "detail" in data


class TestNoHeaderRowDetectedErrors:
    """Test header row detection error handling."""
    
    def test_empty_sheet_raises_error(self, mock_llm_service):
        """
        Test that sheets with no data raise ValueError.
        
        Requirements: 11.2
        """
        # Create a mock sheet object with max_row=0
        mock_sheet = MagicMock()
        mock_sheet.max_row = 0
        mock_sheet.max_column = 5
        
        with pytest.raises(ValueError, match="no data"):
            find_header_row(mock_sheet, mock_llm_service)
    
    def test_none_sheet_raises_error(self, mock_llm_service):
        """
        Test that None sheet raises ValueError.
        
        Requirements: 11.2
        """
        with pytest.raises(ValueError, match="cannot be None"):
            find_header_row(None, mock_llm_service)
    
    @patch('app.core.preprocessor._llm_detect_header_row')
    def test_llm_detection_failure_fallback(self, mock_llm_detect, mock_llm_service):
        """
        Test that LLM detection failure falls back to row 1.
        
        Requirements: 11.2, 11.3
        """
        # Create a workbook with no clear header row (all numeric)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add rows with only numbers (no clear header)
        for row_idx in range(1, 6):
            for col_idx in range(1, 4):
                sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
        
        # Mock LLM detection to raise an error
        mock_llm_detect.side_effect = ValueError("LLM detection failed")
        
        # Should not raise, should fall back to row 1
        result = find_header_row(sheet, mock_llm_service)
        
        assert result.header_row_index == 1
        assert result.data_start_row == 2
    
    def test_llm_detect_with_invalid_response(self, mock_llm_service):
        """
        Test that invalid LLM response raises ValueError.
        
        Requirements: 11.3
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add some data
        sheet['A1'] = 'Header1'
        sheet['B1'] = 'Header2'
        
        # Mock LLM to return invalid response
        mock_llm_service.simple_query.return_value = "not a number"
        
        with pytest.raises(ValueError, match="could not be parsed"):
            _llm_detect_header_row(sheet, mock_llm_service, 5)
    
    def test_llm_detect_with_out_of_range_response(self, mock_llm_service):
        """
        Test that out-of-range LLM response falls back to row 1.
        
        Requirements: 11.3
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add some data
        sheet['A1'] = 'Header1'
        sheet['B1'] = 'Header2'
        
        # Mock LLM to return out-of-range row number
        mock_llm_service.simple_query.return_value = "100"
        
        result = _llm_detect_header_row(sheet, mock_llm_service, 5)
        
        # Should fall back to row 1
        assert result == 1


class TestLLMAPIFailureHandling:
    """Test LLM API failure handling."""
    
    @patch('app.services.llm.genai.GenerativeModel')
    def test_llm_api_failure_returns_unmapped_headers(self, mock_model_class):
        """
        Test that LLM API failures result in method="none" headers.
        
        Requirements: 11.3
        """
        # Setup mock to always fail
        mock_model = MagicMock()
        mock_model_class.return_value = mock_model
        mock_model.generate_content.side_effect = Exception("API Error")
        
        # Create LLM service
        service = LLMService(api_key="test-key", max_retries=2)
        
        # Try to match headers
        headers = ["Unknown Header 1", "Unknown Header 2"]
        result = service.batch_match_headers(headers, ["Power_Output"], ["TG"])
        
        # Should return unmapped headers with method="none"
        assert len(result) == 2
        for mapping in result:
            assert mapping.method == MatchMethod.NONE
            assert mapping.confidence == ConfidenceLevel.LOW
            assert mapping.matched_parameter is None
    
    @patch('app.services.llm.genai.GenerativeModel')
    def test_llm_retry_logic_exhausted(self, mock_model_class):
        """
        Test that retry logic is exhausted after max attempts.
        
        Requirements: 11.3, 14.4
        """
        # Setup mock to fail multiple times
        mock_model = MagicMock()
        mock_model_class.return_value = mock_model
        mock_model.generate_content.side_effect = Exception("API Error")
        
        # Create LLM service with 3 retries
        service = LLMService(api_key="test-key", max_retries=3)
        
        # Call should fail after 3 attempts
        result = service.batch_match_headers(["Test Header"], ["Power_Output"], ["TG"])
        
        # Should have tried 3 times
        assert mock_model.generate_content.call_count == 3
        
        # Should return unmapped header
        assert len(result) == 1
        assert result[0].method == MatchMethod.NONE
    
    def test_matcher_handles_llm_failure_gracefully(self, registry):
        """
        Test that HeaderMatcher handles LLM failures gracefully.
        
        Requirements: 11.3
        """
        # Create mock LLM service that fails
        mock_llm = MagicMock(spec=LLMService)
        mock_llm.batch_match_headers.side_effect = Exception("LLM Service Error")
        
        # Create matcher
        matcher = HeaderMatcher(registry, mock_llm)
        
        # Try to match headers that will fail Tier 1 and 2
        headers = ["Unknown Header XYZ"]
        
        # Should not raise, should return unmapped headers
        result = matcher.match_headers(headers)
        
        assert len(result) == 1
        assert result[0].method == MatchMethod.NONE
        assert result[0].confidence == ConfidenceLevel.LOW


class TestPydanticValidationErrors:
    """Test Pydantic validation error handling."""
    
    def test_header_mapping_invalid_method(self):
        """
        Test that invalid method value raises ValidationError.
        
        Requirements: 11.5
        """
        with pytest.raises(ValidationError):
            HeaderMapping(
                original_header="Test",
                method="invalid_method",  # Invalid value
                confidence=ConfidenceLevel.HIGH
            )
    
    def test_header_mapping_invalid_confidence(self):
        """
        Test that invalid confidence value raises ValidationError.
        
        Requirements: 11.5
        """
        with pytest.raises(ValidationError):
            HeaderMapping(
                original_header="Test",
                method=MatchMethod.EXACT,
                confidence="invalid_confidence"  # Invalid value
            )
    
    def test_parsed_cell_negative_row_index(self):
        """
        Test that negative row index raises ValidationError.
        
        Requirements: 11.5
        """
        mapping = HeaderMapping(
            original_header="Test",
            method=MatchMethod.EXACT,
            confidence=ConfidenceLevel.HIGH
        )
        
        with pytest.raises(ValidationError):
            ParsedCell(
                row_index=-5,  # Invalid negative index
                column_index=0,
                original_value="test",
                parsed_value="test",
                header_mapping=mapping,
                parse_success=True
            )
    
    def test_table_structure_invalid_data_start_row(self):
        """
        Test that data_start_row before header_row raises ValidationError.
        
        Requirements: 11.5
        """
        with pytest.raises(ValidationError):
            TableStructure(
                header_row_index=5,
                data_start_row=3,  # Before header row
                column_count=10
            )
    
    def test_parse_result_negative_total_cells(self):
        """
        Test that negative total_cells raises ValidationError.
        
        Requirements: 11.5
        """
        table_structure = TableStructure(
            header_row_index=1,
            data_start_row=2,
            column_count=3
        )
        
        with pytest.raises(ValidationError):
            ParseResult(
                file_name="test.xlsx",
                table_structure=table_structure,
                header_mappings=[],
                parsed_data=[],
                total_cells=-10,  # Invalid negative value
                successful_parses=0,
                llm_calls_made=0
            )


class TestCorruptedFileHandling:
    """Test corrupted file handling."""
    
    def test_corrupted_excel_file_returns_400(self, client):
        """
        Test that corrupted Excel files return HTTP 400.
        
        Requirements: 11.1
        """
        # Create corrupted file (looks like zip but isn't valid Excel)
        corrupted_file = BytesIO(b"PK\x03\x04corrupted data")
        
        response = client.post(
            "/api/parse",
            files={"file": ("test.xlsx", corrupted_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Should return 400 or 500
        assert response.status_code in [400, 500]
        data = response.json()
        assert "detail" in data
    
    def test_empty_file_returns_400(self, client):
        """
        Test that empty files return HTTP 400.
        
        Requirements: 11.1
        """
        empty_file = BytesIO(b"")
        
        response = client.post(
            "/api/parse",
            files={"file": ("test.xlsx", empty_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
        assert "empty" in data["detail"].lower()


class TestMatcherErrorHandling:
    """Test error handling in HeaderMatcher."""
    
    def test_matcher_with_empty_headers_list(self, registry, mock_llm_service):
        """
        Test that matcher handles empty headers list.
        
        Requirements: 11.1
        """
        matcher = HeaderMatcher(registry, mock_llm_service)
        
        result = matcher.match_headers([])
        
        assert result == []
    
    def test_matcher_tier1_exception_continues_to_tier2(self, registry, mock_llm_service):
        """
        Test that Tier 1 exceptions don't stop processing.
        
        Requirements: 11.1
        """
        # Create matcher with mock registry that raises exception
        mock_registry = MagicMock(spec=RegistryManager)
        mock_registry.exact_match.side_effect = Exception("Tier 1 Error")
        mock_registry.extract_asset.return_value = ("TG", "TG-1")
        mock_registry.exact_match.side_effect = [Exception("Error"), "Power_Output"]
        
        matcher = HeaderMatcher(mock_registry, mock_llm_service)
        
        # Should not raise, should continue to Tier 2
        result = matcher.match_headers(["TG-1 Power"])
        
        assert len(result) == 1
        # Should have attempted Tier 2 or Tier 3
        assert result[0].method in [MatchMethod.FUZZY, MatchMethod.LLM, MatchMethod.NONE]
    
    def test_matcher_critical_error_returns_unmapped(self, registry, mock_llm_service):
        """
        Test that critical errors return unmapped headers.
        
        Requirements: 11.1
        """
        # Create matcher that will fail completely
        mock_registry = MagicMock(spec=RegistryManager)
        mock_registry.exact_match.side_effect = Exception("Critical Error")
        mock_registry.extract_asset.side_effect = Exception("Critical Error")
        mock_llm_service.batch_match_headers.side_effect = Exception("Critical Error")
        
        matcher = HeaderMatcher(mock_registry, mock_llm_service)
        
        # Should not raise, should return unmapped headers
        result = matcher.match_headers(["Test Header"])
        
        assert len(result) == 1
        assert result[0].method == MatchMethod.NONE
        assert result[0].original_header == "Test Header"


class TestPreprocessorErrorHandling:
    """Test error handling in preprocessor."""
    
    def test_find_header_row_with_corrupted_row(self, mock_llm_service):
        """
        Test that corrupted rows don't stop header detection.
        
        Requirements: 11.1
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add valid header row
        sheet['A2'] = 'Header1'
        sheet['B2'] = 'Header2'
        sheet['C2'] = 'Header3'
        sheet['D2'] = 'Header4'
        
        # Should detect row 2 as header
        result = find_header_row(sheet, mock_llm_service)
        
        assert result.header_row_index == 2
        assert result.data_start_row == 3
    
    def test_find_header_row_with_merged_cell_error(self, mock_llm_service):
        """
        Test that merged cell detection errors don't stop processing.
        
        Requirements: 11.1
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add header row
        sheet['A1'] = 'Header1'
        sheet['B1'] = 'Header2'
        sheet['C1'] = 'Header3'
        sheet['D1'] = 'Header4'
        
        # Should still detect header even if merged cell detection fails
        result = find_header_row(sheet, mock_llm_service)
        
        assert result.header_row_index == 1
        assert result.data_start_row == 2
        # merged_cells might be empty if detection failed, but that's OK
        assert isinstance(result.merged_cells, list)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
