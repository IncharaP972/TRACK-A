"""
Unit tests for the Excel preprocessor.
Tests specific examples and edge cases for header row detection and merged cell handling.
"""

import pytest
from openpyxl import Workbook
from unittest.mock import Mock, MagicMock
from app.core.preprocessor import find_header_row, _llm_detect_header_row
from app.schema.models import TableStructure
from app.services.llm import LLMService


class TestFindHeaderRow:
    """Test find_header_row function with specific examples"""
    
    def test_single_clear_header_row(self):
        """Test detection of a single clear header row"""
        wb = Workbook()
        sheet = wb.active
        
        # Add header row at row 1 with 5 string values
        headers = ["Power_Output", "Temperature", "Efficiency", "Status", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Add data rows
        for row_idx in range(2, 5):
            for col_idx in range(1, 6):
                sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
        
        # Detect header row
        result = find_header_row(sheet)
        
        assert isinstance(result, TableStructure)
        assert result.header_row_index == 1
        assert result.data_start_row == 2
        assert result.column_count == 5
        assert len(result.merged_cells) == 0
    
    def test_header_row_not_at_top(self):
        """Test detection when header row is not at row 1"""
        wb = Workbook()
        sheet = wb.active
        
        # Add title rows with few strings
        sheet.cell(row=1, column=1, value="Report Title")
        sheet.cell(row=2, column=1, value="Date: 2024-01-01")
        
        # Add header row at row 3 with many strings
        headers = ["Asset", "Parameter", "Value", "Unit", "Timestamp"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=3, column=col_idx, value=header)
        
        # Add data rows
        for row_idx in range(4, 7):
            for col_idx in range(1, 6):
                sheet.cell(row=row_idx, column=col_idx, value=row_idx * 10)
        
        # Detect header row
        result = find_header_row(sheet)
        
        assert result.header_row_index == 3
        assert result.data_start_row == 4
        assert result.column_count == 5
    
    def test_multiple_candidate_rows_selects_most_strings(self):
        """Test that when multiple rows have >3 strings, the one with most is selected"""
        wb = Workbook()
        sheet = wb.active
        
        # Row 1: 4 strings
        for col_idx in range(1, 5):
            sheet.cell(row=1, column=col_idx, value=f"String_{col_idx}")
        
        # Row 2: 6 strings (should be selected)
        for col_idx in range(1, 7):
            sheet.cell(row=2, column=col_idx, value=f"Header_{col_idx}")
        
        # Row 3: 5 strings
        for col_idx in range(1, 6):
            sheet.cell(row=3, column=col_idx, value=f"Text_{col_idx}")
        
        # Detect header row
        result = find_header_row(sheet)
        
        assert result.header_row_index == 2
        assert result.data_start_row == 3
    
    def test_no_clear_candidate_defaults_to_row_1(self):
        """Test that when no clear candidate exists, defaults to row 1"""
        wb = Workbook()
        sheet = wb.active
        
        # Add rows with only numeric data (no strings)
        for row_idx in range(1, 5):
            for col_idx in range(1, 6):
                sheet.cell(row=row_idx, column=col_idx, value=row_idx * col_idx)
        
        # Detect header row (without LLM service)
        result = find_header_row(sheet, llm_service=None)
        
        assert result.header_row_index == 1
        assert result.data_start_row == 2
    
    def test_merged_cells_detection(self):
        """Test detection of merged cells"""
        wb = Workbook()
        sheet = wb.active
        
        # Add header row
        for col_idx in range(1, 6):
            sheet.cell(row=1, column=col_idx, value=f"Header_{col_idx}")
        
        # Merge cells A2:B3
        sheet.merge_cells(start_row=2, start_column=1, end_row=3, end_column=2)
        
        # Merge cells D2:E2
        sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
        
        # Detect table structure
        result = find_header_row(sheet)
        
        assert len(result.merged_cells) == 2
        
        # Check that both expected merged ranges are present (order may vary)
        expected_ranges = [
            {"min_row": 2, "max_row": 3, "min_col": 1, "max_col": 2},
            {"min_row": 2, "max_row": 2, "min_col": 4, "max_col": 5}
        ]
        
        for expected in expected_ranges:
            found = any(
                merged["min_row"] == expected["min_row"] and
                merged["max_row"] == expected["max_row"] and
                merged["min_col"] == expected["min_col"] and
                merged["max_col"] == expected["max_col"]
                for merged in result.merged_cells
            )
            assert found, f"Expected merged range {expected} not found in {result.merged_cells}"
    
    def test_no_merged_cells(self):
        """Test that merged_cells is empty when no cells are merged"""
        wb = Workbook()
        sheet = wb.active
        
        # Add header row
        for col_idx in range(1, 4):
            sheet.cell(row=1, column=col_idx, value=f"Col_{col_idx}")
        
        # Detect table structure
        result = find_header_row(sheet)
        
        assert result.merged_cells == []
    
    def test_header_with_mixed_content(self):
        """Test header detection with mixed string and numeric content"""
        wb = Workbook()
        sheet = wb.active
        
        # Row 1: Mix of strings and numbers (only 2 strings)
        sheet.cell(row=1, column=1, value="Title")
        sheet.cell(row=1, column=2, value=123)
        sheet.cell(row=1, column=3, value=456)
        sheet.cell(row=1, column=4, value="Date")
        
        # Row 2: Many strings (should be selected)
        headers = ["Asset", "Parameter", "Value", "Unit", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col_idx, value=header)
        
        # Detect header row
        result = find_header_row(sheet)
        
        assert result.header_row_index == 2
        assert result.data_start_row == 3
    
    def test_empty_cells_in_header_row(self):
        """Test header detection with empty cells in the header row"""
        wb = Workbook()
        sheet = wb.active
        
        # Header row with some empty cells
        sheet.cell(row=1, column=1, value="Col1")
        sheet.cell(row=1, column=2, value="Col2")
        # Column 3 is empty
        sheet.cell(row=1, column=4, value="Col4")
        sheet.cell(row=1, column=5, value="Col5")
        sheet.cell(row=1, column=6, value="Col6")
        
        # Detect header row (still has >3 strings)
        result = find_header_row(sheet)
        
        assert result.header_row_index == 1
        assert result.data_start_row == 2
    
    def test_whitespace_only_strings_not_counted(self):
        """Test that whitespace-only strings are not counted as valid strings"""
        wb = Workbook()
        sheet = wb.active
        
        # Row 1: Whitespace-only strings (should not be counted)
        sheet.cell(row=1, column=1, value="   ")
        sheet.cell(row=1, column=2, value="\t")
        sheet.cell(row=1, column=3, value="")
        
        # Row 2: Valid strings (should be selected)
        headers = ["Header1", "Header2", "Header3", "Header4"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col_idx, value=header)
        
        # Detect header row
        result = find_header_row(sheet)
        
        assert result.header_row_index == 2
        assert result.data_start_row == 3
    
    def test_column_count_matches_sheet_max_column(self):
        """Test that column_count matches the sheet's max_column"""
        wb = Workbook()
        sheet = wb.active
        
        # Add header row with 8 columns
        for col_idx in range(1, 9):
            sheet.cell(row=1, column=col_idx, value=f"Col_{col_idx}")
        
        # Detect table structure
        result = find_header_row(sheet)
        
        assert result.column_count == 8
        assert result.column_count == sheet.max_column


class TestLLMDetectHeaderRow:
    """Test _llm_detect_header_row function with LLM service"""
    
    def test_llm_detection_with_valid_response(self):
        """Test LLM header detection with valid integer response"""
        wb = Workbook()
        sheet = wb.active
        
        # Add some rows
        for row_idx in range(1, 6):
            for col_idx in range(1, 4):
                sheet.cell(row=row_idx, column=col_idx, value=f"R{row_idx}C{col_idx}")
        
        # Mock LLM service
        mock_llm = Mock(spec=LLMService)
        mock_llm.simple_query.return_value = "3"
        
        # Call LLM detection
        result = _llm_detect_header_row(sheet, mock_llm, max_rows=5)
        
        assert result == 3
        assert mock_llm.simple_query.called
    
    def test_llm_detection_with_whitespace_in_response(self):
        """Test LLM detection handles whitespace in response"""
        wb = Workbook()
        sheet = wb.active
        
        # Add some rows
        for row_idx in range(1, 4):
            for col_idx in range(1, 3):
                sheet.cell(row=row_idx, column=col_idx, value=f"Data_{row_idx}_{col_idx}")
        
        # Mock LLM service with whitespace in response
        mock_llm = Mock(spec=LLMService)
        mock_llm.simple_query.return_value = "  2  \n"
        
        # Call LLM detection
        result = _llm_detect_header_row(sheet, mock_llm, max_rows=3)
        
        assert result == 2
    
    def test_llm_detection_with_invalid_response_defaults_to_1(self):
        """Test LLM detection defaults to row 1 when response is invalid"""
        wb = Workbook()
        sheet = wb.active
        
        # Add some rows
        for row_idx in range(1, 4):
            sheet.cell(row=row_idx, column=1, value=f"Row_{row_idx}")
        
        # Mock LLM service with invalid response
        mock_llm = Mock(spec=LLMService)
        mock_llm.simple_query.return_value = "not a number"
        
        # Call LLM detection - should raise ValueError
        with pytest.raises(ValueError, match="could not be parsed"):
            _llm_detect_header_row(sheet, mock_llm, max_rows=3)
    
    def test_llm_detection_with_empty_response_defaults_to_1(self):
        """Test LLM detection raises ValueError when response is empty"""
        wb = Workbook()
        sheet = wb.active
        
        # Add some rows
        for row_idx in range(1, 3):
            sheet.cell(row=row_idx, column=1, value=f"Data_{row_idx}")
        
        # Mock LLM service with empty response
        mock_llm = Mock(spec=LLMService)
        mock_llm.simple_query.return_value = ""
        
        # Call LLM detection - should raise ValueError
        with pytest.raises(ValueError, match="empty response"):
            _llm_detect_header_row(sheet, mock_llm, max_rows=2)
    
    def test_llm_detection_prompt_includes_rows(self):
        """Test that LLM detection prompt includes the specified rows"""
        wb = Workbook()
        sheet = wb.active
        
        # Add specific data
        sheet.cell(row=1, column=1, value="Title")
        sheet.cell(row=2, column=1, value="Header1")
        sheet.cell(row=2, column=2, value="Header2")
        sheet.cell(row=3, column=1, value=123)
        
        # Mock LLM service
        mock_llm = Mock(spec=LLMService)
        mock_llm.simple_query.return_value = "2"
        
        # Call LLM detection
        result = _llm_detect_header_row(sheet, mock_llm, max_rows=3)
        
        # Verify the prompt was called
        assert mock_llm.simple_query.called
        call_args = mock_llm.simple_query.call_args[0][0]
        
        # Check that the prompt contains row information
        assert "Row 1:" in call_args
        assert "Row 2:" in call_args
        assert "Row 3:" in call_args
        assert "Title" in call_args
        assert "Header1" in call_args


class TestTableStructureValidation:
    """Test TableStructure model validation"""
    
    def test_table_structure_creation(self):
        """Test creating a valid TableStructure"""
        structure = TableStructure(
            header_row_index=2,
            data_start_row=3,
            column_count=5,
            merged_cells=[]
        )
        
        assert structure.header_row_index == 2
        assert structure.data_start_row == 3
        assert structure.column_count == 5
        assert structure.merged_cells == []
    
    def test_table_structure_with_merged_cells(self):
        """Test TableStructure with merged cell data"""
        merged_cells = [
            {"min_row": 1, "max_row": 2, "min_col": 1, "max_col": 2},
            {"min_row": 3, "max_row": 3, "min_col": 4, "max_col": 6}
        ]
        
        structure = TableStructure(
            header_row_index=1,
            data_start_row=2,
            column_count=10,
            merged_cells=merged_cells
        )
        
        assert len(structure.merged_cells) == 2
        assert structure.merged_cells[0]["min_row"] == 1
        assert structure.merged_cells[1]["max_col"] == 6
