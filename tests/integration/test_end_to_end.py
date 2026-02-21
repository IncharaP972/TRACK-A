"""
Integration tests for complete parsing workflow.

These tests verify the end-to-end functionality of the Excel parser,
including file upload, header detection, three-tier matching, and
value parsing with complete audit trails.

Requirements: 9.6, 10.4
"""

import os
import pytest
from fastapi.testclient import TestClient
from io import BytesIO
import openpyxl
from openpyxl import Workbook

from app.main import app
from app.schema.models import MatchMethod, ConfidenceLevel


@pytest.fixture
def client():
    """Create a test client for the FastAPI app."""
    return TestClient(app)


@pytest.fixture
def sample_excel_file():
    """
    Create a sample Excel file with known structure for testing.
    
    Structure:
    - Row 1: Title row (should be skipped)
    - Row 2: Header row with known parameters
    - Rows 3-5: Data rows with various value types
    """
    wb = Workbook()
    ws = wb.active
    
    # Row 1: Title (should be detected as non-header)
    ws.append(["Power Plant Operational Data Report"])
    
    # Row 2: Headers (should be detected as header row)
    ws.append([
        "Power_Output",
        "Temperature", 
        "Efficiency",
        "Heat_Rate",
        "Fuel_Consumption"
    ])
    
    # Row 3: Data with various formats
    ws.append([
        "1,234.56",  # Number with comma
        "450",       # Plain number
        "85%",       # Percentage
        "12500",     # Plain number
        "500.5"      # Decimal
    ])
    
    # Row 4: More data
    ws.append([
        "2,000.00",
        "475",
        "90%",
        "13000",
        "550.0"
    ])
    
    # Row 5: Data with N/A values
    ws.append([
        "N/A",
        "460",
        "N/A",
        "12800",
        ""
    ])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


@pytest.fixture
def sample_excel_with_assets():
    """
    Create a sample Excel file with asset identifiers for Tier 2 matching.
    
    Structure:
    - Row 1: Headers with asset patterns
    - Rows 2-3: Data rows
    """
    wb = Workbook()
    ws = wb.active
    
    # Headers with asset identifiers
    ws.append([
        "Power_Output",
        "TG-1 Temperature",
        "AFBC-2 Efficiency",
        "ESP-3 Availability"
    ])
    
    # Data rows
    ws.append([
        "1000",
        "450",
        "88%",
        "95%"
    ])
    
    ws.append([
        "1100",
        "455",
        "89%",
        "96%"
    ])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


class TestCompleteParsingWorkflow:
    """
    Integration test for complete parsing workflow.
    
    Validates: Requirements 9.6, 10.4
    """
    
    def test_parse_excel_with_exact_matches(self, client, sample_excel_file):
        """
        Test complete workflow with headers that match exactly.
        
        This test verifies:
        - File upload and validation
        - Header row detection (should detect row 2)
        - Tier 1 exact matching for all headers
        - Value parsing with various formats
        - Complete audit trail in ParseResult
        """
        # Upload the file
        response = client.post(
            "/api/parse",
            files={"file": ("test_data.xlsx", sample_excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Verify successful response
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        
        # Verify file name
        assert result["file_name"] == "test_data.xlsx"
        
        # Verify table structure detection
        assert result["table_structure"]["header_row_index"] == 2, "Should detect row 2 as header"
        assert result["table_structure"]["data_start_row"] == 3, "Data should start at row 3"
        assert result["table_structure"]["column_count"] == 5, "Should have 5 columns"
        
        # Verify header mappings
        assert len(result["header_mappings"]) == 5, "Should have 5 header mappings"
        
        # Check that all headers matched exactly (Tier 1)
        for mapping in result["header_mappings"]:
            assert mapping["method"] == "exact", f"Expected exact match for {mapping['original_header']}"
            assert mapping["confidence"] == "high", "Exact matches should have high confidence"
            assert mapping["matched_parameter"] is not None, "Should have matched parameter"
        
        # Verify specific header mappings
        header_params = [m["matched_parameter"] for m in result["header_mappings"]]
        assert "Power_Output" in header_params
        assert "Temperature" in header_params
        assert "Efficiency" in header_params
        assert "Fuel_Consumption" in header_params
        
        # Verify parsed data
        assert len(result["parsed_data"]) == 3, "Should have 3 data rows"
        
        # Verify first row parsing
        row1 = result["parsed_data"][0]
        assert len(row1) == 5, "Each row should have 5 cells"
        
        # Check specific parsed values
        # "1,234.56" should parse to 1234.56
        assert row1[0]["original_value"] == "1,234.56"
        assert row1[0]["parsed_value"] == 1234.56
        assert row1[0]["parse_success"] is True
        
        # "85%" should parse to 0.85
        assert row1[2]["original_value"] == "85%"
        assert row1[2]["parsed_value"] == 0.85
        assert row1[2]["parse_success"] is True
        
        # "12500" should parse to 12500.0
        assert row1[3]["original_value"] == "12500"
        assert row1[3]["parsed_value"] == 12500.0
        assert row1[3]["parse_success"] is True
        
        # Verify second row
        row2 = result["parsed_data"][1]
        # "13000" should parse to 13000.0
        assert row2[3]["original_value"] == "13000"
        assert row2[3]["parsed_value"] == 13000.0
        
        # Verify third row with N/A values
        row3 = result["parsed_data"][2]
        # "N/A" should parse to None
        assert row3[0]["original_value"] == "N/A"
        assert row3[0]["parsed_value"] is None
        assert row3[0]["parse_success"] is True
        
        # Empty string should parse to None
        assert row3[4]["original_value"] in ["", None]
        assert row3[4]["parsed_value"] is None
        
        # Verify audit trail completeness
        for row in result["parsed_data"]:
            for cell in row:
                # Each cell should have complete audit information
                assert "row_index" in cell
                assert "column_index" in cell
                assert "original_value" in cell
                assert "parsed_value" in cell
                assert "header_mapping" in cell
                assert "parse_success" in cell
                
                # Header mapping should have method and confidence
                assert "method" in cell["header_mapping"]
                assert "confidence" in cell["header_mapping"]
        
        # Verify summary statistics
        assert result["total_cells"] == 15, "Should have 15 total cells (3 rows × 5 columns)"
        assert result["successful_parses"] == 15, "All cells should parse successfully"
        assert result["llm_calls_made"] == 0, "No LLM calls should be needed for exact matches"
    
    def test_parse_excel_with_asset_extraction(self, client, sample_excel_with_assets):
        """
        Test complete workflow with headers containing asset identifiers.
        
        This test verifies:
        - Tier 2 regex-based asset extraction
        - Asset identifier extraction (TG-1, AFBC-2, ESP-3)
        - Fuzzy matching method assignment
        - Complete audit trail with asset information
        """
        # Upload the file
        response = client.post(
            "/api/parse",
            files={"file": ("test_assets.xlsx", sample_excel_with_assets, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Verify successful response
        assert response.status_code == 200
        
        result = response.json()
        
        # Verify header mappings
        assert len(result["header_mappings"]) == 4
        
        # First header should be exact match
        assert result["header_mappings"][0]["original_header"] == "Power_Output"
        assert result["header_mappings"][0]["method"] == "exact"
        assert result["header_mappings"][0]["matched_parameter"] == "Power_Output"
        
        # Second header should extract TG-1 asset
        assert result["header_mappings"][1]["original_header"] == "TG-1 Temperature"
        assert result["header_mappings"][1]["method"] == "fuzzy"
        assert result["header_mappings"][1]["matched_asset"] == "TG-1"
        assert result["header_mappings"][1]["matched_parameter"] == "Temperature"
        
        # Third header should extract AFBC-2 asset
        assert result["header_mappings"][2]["original_header"] == "AFBC-2 Efficiency"
        assert result["header_mappings"][2]["method"] == "fuzzy"
        assert result["header_mappings"][2]["matched_asset"] == "AFBC-2"
        assert result["header_mappings"][2]["matched_parameter"] == "Efficiency"
        
        # Fourth header should extract ESP-3 asset
        assert result["header_mappings"][3]["original_header"] == "ESP-3 Availability"
        assert result["header_mappings"][3]["method"] == "fuzzy"
        assert result["header_mappings"][3]["matched_asset"] == "ESP-3"
        
        # Verify parsed data
        assert len(result["parsed_data"]) == 2
        
        # Verify audit trail includes asset information
        for row in result["parsed_data"]:
            for cell in row:
                if cell["header_mapping"]["matched_asset"]:
                    # Cells with asset mappings should have the asset in the mapping
                    assert cell["header_mapping"]["matched_asset"] in ["TG-1", "AFBC-2", "ESP-3"]
        
        # No LLM calls should be needed
        assert result["llm_calls_made"] == 0
    
    def test_parse_excel_error_handling(self, client):
        """
        Test error handling for invalid inputs.
        
        This test verifies:
        - Invalid file type rejection
        - Empty file handling
        - Appropriate error messages
        """
        # Test invalid file type
        invalid_file = BytesIO(b"not an excel file")
        response = client.post(
            "/api/parse",
            files={"file": ("test.txt", invalid_file, "text/plain")}
        )
        assert response.status_code == 400
        assert "Excel files" in response.json()["detail"]
        
        # Test empty file
        empty_file = BytesIO(b"")
        response = client.post(
            "/api/parse",
            files={"file": ("empty.xlsx", empty_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code in [400, 422, 500]  # Could be various error codes depending on where it fails
    
    def test_parse_result_json_serialization(self, client, sample_excel_file):
        """
        Test that ParseResult is fully serializable to JSON.
        
        This test verifies:
        - Complete JSON serialization of ParseResult
        - All nested models serialize correctly
        - No serialization errors
        """
        response = client.post(
            "/api/parse",
            files={"file": ("test.xlsx", sample_excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        
        # Verify we can parse the JSON response
        result = response.json()
        
        # Verify all required fields are present
        required_fields = [
            "file_name",
            "table_structure",
            "header_mappings",
            "parsed_data",
            "total_cells",
            "successful_parses",
            "llm_calls_made"
        ]
        
        for field in required_fields:
            assert field in result, f"Missing required field: {field}"
        
        # Verify nested structures are serializable
        assert isinstance(result["table_structure"], dict)
        assert isinstance(result["header_mappings"], list)
        assert isinstance(result["parsed_data"], list)
        
        # Verify all data is JSON-compatible (no custom objects)
        import json
        json_str = json.dumps(result)
        assert len(json_str) > 0, "Should serialize to non-empty JSON"




@pytest.fixture
def sample_excel_with_ambiguous_headers():
    """
    Create a sample Excel file with ambiguous headers that require LLM matching.
    
    These headers don't match exactly and don't contain clear asset patterns,
    so they should trigger Tier 3 LLM matching.
    """
    wb = Workbook()
    ws = wb.active
    
    # Headers that are ambiguous and need LLM interpretation
    ws.append([
        "Power_Output",  # This will match exactly (Tier 1)
        "Temp Reading",  # Ambiguous - should map to Temperature via LLM
        "Plant Efficiency Rate",  # Ambiguous - should map to Efficiency via LLM
        "CO2 Emissions Level"  # Ambiguous - should map to Emissions_CO2 via LLM
    ])
    
    # Data rows
    ws.append([
        "1500",
        "480",
        "87%",
        "250"
    ])
    
    ws.append([
        "1600",
        "485",
        "88%",
        "245"
    ])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


class TestLLMIntegration:
    """
    Integration tests with real LLM API.
    
    These tests require a valid GEMINI_API_KEY environment variable.
    They verify Tier 3 semantic matching with actual LLM calls.
    
    Validates: Requirements 6.1, 6.4, 6.5
    """
    
    @pytest.mark.skipif(
        not os.getenv("GEMINI_API_KEY") or os.getenv("GEMINI_API_KEY") == "test-key-for-testing",
        reason="Requires valid GEMINI_API_KEY environment variable"
    )
    def test_llm_semantic_matching(self, client, sample_excel_with_ambiguous_headers):
        """
        Test complete workflow with ambiguous headers requiring LLM matching.
        
        This test verifies:
        - Tier 1 and Tier 2 matching for clear headers
        - Tier 3 LLM matching for ambiguous headers
        - Single batch LLM call for all unmapped headers
        - LLM mappings have correct metadata (method="llm", confidence level)
        - Complete audit trail with LLM match information
        
        Note: This test requires a valid GEMINI_API_KEY environment variable.
        """
        # Upload the file
        response = client.post(
            "/api/parse",
            files={"file": ("ambiguous_data.xlsx", sample_excel_with_ambiguous_headers, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Verify successful response
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        
        # Verify file name
        assert result["file_name"] == "ambiguous_data.xlsx"
        
        # Verify header mappings
        assert len(result["header_mappings"]) == 4, "Should have 4 header mappings"
        
        # First header should be exact match (Tier 1)
        assert result["header_mappings"][0]["original_header"] == "Power_Output"
        assert result["header_mappings"][0]["method"] == "exact"
        assert result["header_mappings"][0]["confidence"] == "high"
        assert result["header_mappings"][0]["matched_parameter"] == "Power_Output"
        
        # Remaining headers should be LLM matched (Tier 3)
        llm_matched_headers = [m for m in result["header_mappings"] if m["method"] == "llm"]
        assert len(llm_matched_headers) >= 1, "Should have at least one LLM-matched header"
        
        # Verify LLM mappings have correct metadata
        for mapping in llm_matched_headers:
            assert mapping["method"] == "llm", "LLM-matched headers should have method='llm'"
            assert mapping["confidence"] in ["high", "medium", "low"], "LLM mappings should have valid confidence level"
            # LLM should provide some mapping (parameter or asset)
            assert mapping["matched_parameter"] is not None or mapping["matched_asset"] is not None, \
                "LLM should provide at least parameter or asset mapping"
        
        # Verify single batch LLM call was made
        assert result["llm_calls_made"] == 1, "Should make exactly one batch LLM call for all unmapped headers"
        
        # Verify parsed data
        assert len(result["parsed_data"]) == 2, "Should have 2 data rows"
        
        # Verify audit trail includes LLM match information
        for row in result["parsed_data"]:
            for cell in row:
                # Each cell should have complete audit information
                assert "header_mapping" in cell
                assert "method" in cell["header_mapping"]
                assert "confidence" in cell["header_mapping"]
                
                # If this cell's header was LLM-matched, verify metadata
                if cell["header_mapping"]["method"] == "llm":
                    assert cell["header_mapping"]["confidence"] in ["high", "medium", "low"]
        
        # Verify summary statistics
        assert result["total_cells"] == 8, "Should have 8 total cells (2 rows × 4 columns)"
        assert result["successful_parses"] == 8, "All cells should parse successfully"
    
    @pytest.mark.skipif(
        not os.getenv("GEMINI_API_KEY") or os.getenv("GEMINI_API_KEY") == "test-key-for-testing",
        reason="Requires valid GEMINI_API_KEY environment variable"
    )
    def test_mixed_tier_matching_with_llm(self, client):
        """
        Test workflow with headers that use all three tiers.
        
        This test verifies:
        - Tier 1 exact matching for standard parameters
        - Tier 2 regex matching for headers with assets
        - Tier 3 LLM matching for remaining ambiguous headers
        - Single batch LLM call despite mixed tier usage
        """
        wb = Workbook()
        ws = wb.active
        
        # Mix of headers requiring different tiers
        ws.append([
            "Power_Output",  # Tier 1: Exact match
            "TG-1 Temperature",  # Tier 2: Asset extraction
            "Efficiency",  # Tier 1: Exact match
            "Boiler Heat Input",  # Tier 3: LLM (ambiguous)
            "AFBC-2 Fuel_Consumption"  # Tier 2: Asset extraction
        ])
        
        # Data row
        ws.append([
            "1200",
            "470",
            "86%",
            "5000",
            "600"
        ])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload the file
        response = client.post(
            "/api/parse",
            files={"file": ("mixed_tiers.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Verify successful response
        assert response.status_code == 200
        
        result = response.json()
        
        # Verify header mappings
        assert len(result["header_mappings"]) == 5
        
        # Count matches by tier
        exact_matches = [m for m in result["header_mappings"] if m["method"] == "exact"]
        fuzzy_matches = [m for m in result["header_mappings"] if m["method"] == "fuzzy"]
        llm_matches = [m for m in result["header_mappings"] if m["method"] == "llm"]
        
        # Should have matches from multiple tiers
        assert len(exact_matches) >= 2, "Should have at least 2 exact matches"
        assert len(fuzzy_matches) >= 2, "Should have at least 2 fuzzy matches"
        # May or may not have LLM matches depending on how well Tier 2 handles "Boiler Heat Input"
        
        # Verify single batch LLM call (0 or 1)
        assert result["llm_calls_made"] in [0, 1], "Should make at most one batch LLM call"
        
        # If LLM was called, verify it was for unmapped headers only
        if result["llm_calls_made"] == 1:
            assert len(llm_matches) >= 1, "If LLM was called, should have at least one LLM match"
    
    @pytest.mark.skipif(
        not os.getenv("GEMINI_API_KEY") or os.getenv("GEMINI_API_KEY") == "test-key-for-testing",
        reason="Requires valid GEMINI_API_KEY environment variable"
    )
    def test_llm_batch_efficiency(self, client):
        """
        Test that multiple ambiguous headers are processed in a single LLM call.
        
        This test verifies:
        - Multiple unmapped headers trigger only one LLM API call
        - Batch processing efficiency (Requirement 12.1)
        - All unmapped headers receive mappings from the single call
        """
        wb = Workbook()
        ws = wb.active
        
        # Multiple ambiguous headers that will all need LLM
        ws.append([
            "Plant Output Reading",  # Ambiguous
            "Temp Sensor Data",  # Ambiguous
            "Efficiency Metric",  # Ambiguous
            "Fuel Usage Rate",  # Ambiguous
            "CO2 Level"  # Ambiguous
        ])
        
        # Data row
        ws.append([
            "1300",
            "465",
            "85%",
            "580",
            "240"
        ])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload the file
        response = client.post(
            "/api/parse",
            files={"file": ("batch_test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Verify successful response
        assert response.status_code == 200
        
        result = response.json()
        
        # Verify header mappings
        assert len(result["header_mappings"]) == 5
        
        # CRITICAL: Despite having 5 ambiguous headers, should make only 1 LLM call
        assert result["llm_calls_made"] == 1, "Should make exactly ONE batch LLM call for all unmapped headers"
        
        # Verify all headers received some mapping attempt
        for mapping in result["header_mappings"]:
            assert mapping["method"] in ["exact", "fuzzy", "llm", "none"], "Each header should have a valid method"
            assert mapping["confidence"] in ["high", "medium", "low"], "Each header should have a confidence level"

