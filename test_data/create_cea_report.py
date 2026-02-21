"""
CEA Daily Coal Report Generator - Real World "Dirty Data" Test

This script creates a realistic CEA Daily Coal Stock report with intentional
data quality issues to test the parser's robustness:
- Mixed date formats
- Units embedded in cell values (MW, k Tonnes, T, %)
- Inconsistent formatting (nil, zero, -, 0)
- Merged header cells
- Regional grouping rows
- Abbreviations and shorthand
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_real_world_cea_excel():
    """
    Reconstructs the CEA Daily Coal Report into a "Dirty Data" format.
    Tests the Agent's ability to handle human errors, inconsistent units, 
    and unstructured cell values.
    """
    import os
    file_path = os.path.join(os.path.dirname(__file__), "cea_coal_report_real.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Coal Stock"
    
    # --- SETUP STYLES ---
    header_font = Font(bold=True, size=10)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    fill_blue = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    
    # --- ROW 1-3: LOGO & TITLES (Noise) ---
    ws.merge_cells('A1:O1')
    ws['A1'] = "CENTRAL ELECTRICITY AUTHORITY (CEA)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_aligned
    
    ws.merge_cells('A2:O2')
    ws['A2'] = "DAILY COAL STOCK - STATUS AS ON 21/02/19"  # Mixed date format
    ws['A2'].alignment = center_aligned
    
    ws.append([])  # Row 3: Spacer
    
    # --- ROW 4-5: COMPLEX HEADERS ---
    ws.merge_cells('A4:A5')
    ws['A4'] = "Region/State"
    
    ws.merge_cells('B4:B5')
    ws['B4'] = "Transport Mode"  # Different from PDF to test mapping
    
    ws.merge_cells('C4:C5')
    ws['C4'] = "Power Station Name"
    
    ws.merge_cells('D4:D5')
    ws['D4'] = "Sec."  # Abbreviation
    
    ws.merge_cells('E4:E5')
    ws['E4'] = "MW Cap."
    
    ws.merge_cells('F4:F5')
    ws['F4'] = "PLF% (JAN-19)"
    
    ws.merge_cells('G4:G5')
    ws['G4'] = "PLF% (APR-JAN)"
    
    ws.merge_cells('H4:H5')
    ws['H4'] = "Norm Stock (Days)"
    
    ws.merge_cells('I4:I5')
    ws['I4'] = "Daily Req. (k Tonnes)"
    
    ws.merge_cells('J4:L4')
    ws['J4'] = "STOCK IN HAND"
    ws['J5'] = "Indig."
    ws['K5'] = "Imp."
    ws['L5'] = "Total Stock"
    
    ws.merge_cells('M4:M5')
    ws['M4'] = "Days"
    
    ws.merge_cells('N4:O5')
    ws['N4'] = "Criticality/Remarks"
    
    # Style headers
    for row in ws.iter_rows(min_row=4, max_row=5, max_col=15):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = border
            cell.fill = fill_blue
    
    # --- ROW 6-15: "DIRTY" DATA ---
    # Purposefully adding units inside cells and inconsistent formats
    data = [
        ["NORTHERN REGION", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Haryana", "RAIL", "INDIRA GANDHI STPP", "JV", "1500", "66.0", "57.9", "25", "14.95", "14.17", "0", "14.17", "1 day", "Coal low"],
        ["Haryana", "RAIL", "PANIPAT TPS", "S", "920", "43.6%", "44.8%", "25", "8.07", "115.86", "nil", "115.86", "14", ""],
        ["Punjab", "RAIL", "GOINDWAL SAHIB TPP", "P", "540", "64.4", "47.9", "30", "6.27", "9.32", "-", "9.32", "1", "Low ACQ"],
        ["Punjab", "RAIL", "TALWANDI SABO TPP", "P", "1980 MW", "63.7", "63.1", "25", "17.81", "241.01 T", "58.39 T", "299.4", "17", ""],
        ["Rajasthan", "RAIL", "KOTA TPS", "S", "1240", "83.9", "75.1", "30", "13.82", "109.23", "0", "109.23", "8", "Normal"],
        ["Uttar Pradesh", "PIT", "ANPARA TPS", "S", "2630", "87.0", "87.3", "15", "35.94", "404.45", "0", "404.45", "11", ""],
        ["Uttar Pradesh", "RAIL", "RIHAND STPS", "C", "3000", "89", "85.5", "15", "33.67 k", "1127.72", "0", "1,127.72", "33", ""],
        ["WESTERN REGION", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Gujarat", "RAIL", "UKAI TPS", "S", "1110", "67.9", "70", "30", "11.09", "324.12", "2.47", "326.59", "29", ""],
    ]
    
    for r_idx, row_data in enumerate(data, start=6):
        ws.append(row_data)
        # Style regional header rows
        if row_data[1] == "":
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    
    wb.save(file_path)
    print(f"✓ Generated: {file_path}")
    print(f"  - Complex merged headers (2 rows)")
    print(f"  - Regional grouping rows")
    print(f"  - Mixed units in cells (MW, %, k, T)")
    print(f"  - Inconsistent null values (nil, zero, -, 0)")
    print(f"  - Real-world CEA coal report structure")


if __name__ == "__main__":
    create_real_world_cea_excel()
