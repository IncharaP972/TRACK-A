"""
Test Data Generator for LatSpace Excel Parser

Generates three test Excel files to demonstrate the three-tier matching strategy:
1. clean_data.xlsx - Baseline with exact matches (Tier 1)
2. messy_data.xlsx - Fuzzy matching test with variations (Tier 2)
3. multi_asset.xlsx - Asset detection test with multiple assets (Tier 2 + Tier 3)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


def adjust_column_widths(ws):
    """Adjust column widths based on content, handling merged cells."""
    from openpyxl.cell.cell import MergedCell
    
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        
        for cell in column_cells:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            
            # Get column letter from first non-merged cell
            if column_letter is None:
                column_letter = cell.column_letter
            
            try:
                if cell.value is not None:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
            except:
                pass
        
        # Only adjust if we found a valid column
        if column_letter:
            adjusted_width = max(max_length + 2, 10)  # Minimum width of 10
            ws.column_dimensions[column_letter].width = adjusted_width


def create_clean_data():
    """
    Generate clean_data.xlsx - Baseline file with exact parameter matches.
    
    This file tests Tier 1 (Exact Matching):
    - All headers match exactly with registry parameters
    - Should result in 100% exact matches with high confidence
    - No LLM calls should be needed
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Operational Data"
    
    # Add title row
    ws.append(["Power Plant Operational Data - January 2024"])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add header row (Row 2) - Exact matches from registry
    headers = [
        "Power_Output",
        "Temperature",
        "Efficiency",
        "Fuel_Consumption",
        "Emissions_CO2",
        "Steam_Pressure"
    ]
    ws.append(headers)
    
    # Style header row
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    data = [
        ["1,234.56", "450", "85%", "500.5", "250", "120"],
        ["1,500.00", "475", "87%", "520.0", "245", "125"],
        ["1,350.75", "460", "86%", "510.2", "248", "122"],
        ["1,425.00", "470", "88%", "515.8", "242", "123"],
        ["1,380.50", "465", "85%", "505.5", "252", "121"],
        ["N/A", "455", "N/A", "498.0", "240", "119"],
        ["1,290.00", "450", "84%", "495.5", "255", "120"],
        ["1,410.25", "468", "87%", "512.0", "244", "124"],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Adjust column widths
    adjust_column_widths(ws)
    
    # Save file
    output_path = os.path.join(os.path.dirname(__file__), "clean_data.xlsx")
    wb.save(output_path)
    print(f"✓ Created: {output_path}")
    print(f"  - Headers: {len(headers)} exact matches")
    print(f"  - Data rows: {len(data)}")
    print(f"  - Expected: 100% Tier 1 (Exact) matching")


def create_messy_data():
    """
    Generate messy_data.xlsx - File with variations and fuzzy matching needs.
    
    This file tests Tier 2 (Regex/Asset Extraction):
    - Headers with asset identifiers (TG-1, AFBC-2, etc.)
    - Some exact matches, some fuzzy matches
    - Tests asset extraction and parameter inference
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plant Data"
    
    # Add title rows
    ws.append(["LatSpace Power Plant"])
    ws.append(["Operational Metrics - Q1 2024"])
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(italic=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add header row (Row 3) - Mix of exact and fuzzy matches
    headers = [
        "Power_Output",           # Exact match
        "TG-1 Temperature",       # Asset extraction + parameter inference
        "AFBC-2 Efficiency",      # Asset extraction + parameter inference
        "ESP-3 Status",           # Asset extraction (no parameter match)
        "Fuel_Consumption",       # Exact match
        "ID-FAN-1 Speed",         # Asset extraction (no parameter match)
        "Emissions_CO2"           # Exact match
    ]
    ws.append(headers)
    
    # Style header row
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows with various formats
    data = [
        ["1,200", "445", "82%", "Running", "480.5", "1200", "235"],
        ["1,350", "450", "85%", "Running", "495.0", "1250", "240"],
        ["1,280", "448", "83%", "Running", "488.2", "1220", "238"],
        ["1,420", "455", "86%", "Running", "502.5", "1280", "242"],
        ["1,310", "449", "84%", "Stopped", "490.0", "0", "236"],
        ["1,390", "453", "85%", "Running", "498.5", "1260", "241"],
        ["N/A", "N/A", "N/A", "Maintenance", "N/A", "N/A", "N/A"],
        ["1,340", "451", "84%", "Running", "492.0", "1240", "239"],
        ["1,405", "454", "86%", "Running", "500.0", "1270", "243"],
        ["1,295", "447", "83%", "Running", "485.5", "1210", "237"],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Adjust column widths
    adjust_column_widths(ws)
    
    # Save file
    output_path = os.path.join(os.path.dirname(__file__), "messy_data.xlsx")
    wb.save(output_path)
    print(f"✓ Created: {output_path}")
    print(f"  - Headers: {len(headers)} (mix of exact and fuzzy)")
    print(f"  - Data rows: {len(data)}")
    print(f"  - Expected: Tier 1 (3 exact) + Tier 2 (4 fuzzy)")


def create_multi_asset():
    """
    Generate multi_asset.xlsx - File with multiple assets and ambiguous headers.
    
    This file tests all three tiers:
    - Some exact matches (Tier 1)
    - Multiple asset identifiers (Tier 2)
    - Ambiguous headers requiring LLM (Tier 3)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Asset Data"
    
    # Add header row (Row 1) - Mix of all three tiers
    headers = [
        "Timestamp",                    # Unmapped (no match)
        "TG-1 Power_Output",           # Asset + exact parameter
        "TG-2 Power_Output",           # Asset + exact parameter
        "AFBC-1 Temperature",          # Asset + exact parameter
        "AFBC-2 Temperature",          # Asset + exact parameter
        "Boiler Heat Input",           # Ambiguous - needs LLM
        "Plant Efficiency Rate",       # Ambiguous - needs LLM
        "CO2 Emissions Level",         # Ambiguous - needs LLM
        "Fuel_Consumption",            # Exact match
        "ESP-1 Availability"           # Asset extraction
    ]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    data = [
        ["2024-01-01 08:00", "600", "650", "440", "445", "5000", "85%", "240", "480", "98%"],
        ["2024-01-01 09:00", "620", "670", "442", "447", "5100", "86%", "242", "490", "99%"],
        ["2024-01-01 10:00", "610", "660", "441", "446", "5050", "85%", "241", "485", "98%"],
        ["2024-01-01 11:00", "630", "680", "443", "448", "5150", "87%", "243", "495", "99%"],
        ["2024-01-01 12:00", "625", "675", "442", "447", "5120", "86%", "242", "492", "98%"],
        ["2024-01-01 13:00", "615", "665", "441", "446", "5080", "86%", "241", "488", "99%"],
        ["2024-01-01 14:00", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"],
        ["2024-01-01 15:00", "605", "655", "440", "445", "5030", "85%", "240", "482", "98%"],
        ["2024-01-01 16:00", "628", "678", "443", "448", "5140", "87%", "243", "497", "99%"],
        ["2024-01-01 17:00", "618", "668", "442", "447", "5090", "86%", "242", "490", "98%"],
        ["2024-01-01 18:00", "612", "662", "441", "446", "5060", "85%", "241", "486", "98%"],
        ["2024-01-01 19:00", "622", "672", "442", "447", "5110", "86%", "242", "491", "99%"],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Adjust column widths
    adjust_column_widths(ws)
    
    # Save file
    output_path = os.path.join(os.path.dirname(__file__), "multi_asset.xlsx")
    wb.save(output_path)
    print(f"✓ Created: {output_path}")
    print(f"  - Headers: {len(headers)} (all three tiers)")
    print(f"  - Data rows: {len(data)}")
    print(f"  - Expected: Tier 1 (1 exact) + Tier 2 (6 fuzzy) + Tier 3 (3 LLM)")


def main():
    """Generate all three test files."""
    print("=" * 60)
    print("LatSpace Excel Parser - Test Data Generator")
    print("=" * 60)
    print()
    
    create_clean_data()
    print()
    create_messy_data()
    print()
    create_multi_asset()
    
    print()
    print("=" * 60)
    print("✓ All test files generated successfully!")
    print("=" * 60)
    print()
    print("Test Files:")
    print("  1. clean_data.xlsx    - Baseline (100% Tier 1 exact matches)")
    print("  2. messy_data.xlsx    - Fuzzy matching (Tier 1 + Tier 2)")
    print("  3. multi_asset.xlsx   - Multi-tier (Tier 1 + Tier 2 + Tier 3)")
    print()
    print("Upload these files to the frontend to see the three-tier")
    print("matching strategy in action!")
    print()


if __name__ == "__main__":
    main()
