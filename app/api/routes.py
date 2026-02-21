"""
FastAPI routes for Excel file parsing.

This module provides the REST API endpoints for uploading and parsing
Excel files. The main endpoint accepts Excel files, processes them through
the three-tier matching strategy, and returns complete parsing results
with full audit trails.

Requirements: 10.1, 10.4, 10.5, 11.1, 11.2
"""

from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from pydantic import ValidationError
import openpyxl
from io import BytesIO
import logging
import os
from typing import List

from app.core.matcher import HeaderMatcher
from app.core.parser import parse_row
from app.core.preprocessor import find_header_row
from app.schema.models import ParseResult, ParsedCell
from app.registry.data import RegistryManager
from app.services.llm import LLMService

# Configure logging
logger = logging.getLogger(__name__)

# Create router
router = APIRouter()

# Initialize services (in production, use dependency injection)
# For now, we'll initialize them as module-level singletons
_registry = None
_llm_service = None
_matcher = None


def get_registry() -> RegistryManager:
    """Get or create the registry manager singleton."""
    global _registry
    if _registry is None:
        _registry = RegistryManager()
        logger.info("RegistryManager initialized")
    return _registry


def get_llm_service() -> LLMService:
    """Get or create the LLM service singleton."""
    global _llm_service
    if _llm_service is None:
        # Get API key from environment variable
        api_key = os.getenv("GEMINI_API_KEY", "test-key-for-testing")
        _llm_service = LLMService(api_key=api_key)
        logger.info("LLMService initialized")
    return _llm_service


def get_matcher() -> HeaderMatcher:
    """Get or create the header matcher singleton."""
    global _matcher
    if _matcher is None:
        _matcher = HeaderMatcher(get_registry(), get_llm_service())
        logger.info("HeaderMatcher initialized")
    return _matcher


@router.post("/parse", response_model=ParseResult)
async def parse_excel(file: UploadFile = File(...)):
    """
    Upload and parse an Excel file.
    
    This endpoint accepts an Excel file (.xlsx or .xls), processes it through
    the three-tier header matching strategy, and returns complete parsing results
    with full audit trails.
    
    Args:
        file: Uploaded Excel file
        
    Returns:
        ParseResult with complete parsing information and audit trail
        
    Raises:
        HTTPException 400: Invalid file format (not .xlsx or .xls)
        HTTPException 422: No header row detected or table structure issues
        HTTPException 500: Internal processing error
    """
    logger.info(f"Received file upload: {file.filename}")
    
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        logger.warning(f"Invalid file type: {file.filename}")
        raise HTTPException(
            status_code=400, 
            detail="Only Excel files (.xlsx, .xls) are supported"
        )
    
    try:
        # Load Excel file
        logger.debug("Reading file contents")
        contents = await file.read()
        
        if not contents:
            raise HTTPException(status_code=400, detail="Empty file uploaded")
        
        logger.debug("Loading workbook with openpyxl")
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        
        if sheet is None:
            raise HTTPException(status_code=422, detail="No active sheet found in workbook")
        
        logger.info(f"Workbook loaded: {sheet.max_row} rows, {sheet.max_column} columns")
        
        # Step 1: Detect table structure
        logger.debug("Step 1: Detecting table structure")
        try:
            table_structure = find_header_row(sheet, get_llm_service())
        except ValueError as e:
            logger.error(f"Header row detection failed: {e}")
            raise HTTPException(
                status_code=422,
                detail=f"Failed to detect header row: {str(e)}"
            )
        
        if table_structure.header_row_index < 1:
            raise HTTPException(
                status_code=422, 
                detail="No header row detected in the Excel file"
            )
        
        logger.info(f"Table structure detected: header at row {table_structure.header_row_index}")
        
        # Step 2: Extract headers
        logger.debug("Step 2: Extracting headers")
        header_row = list(sheet.iter_rows(
            min_row=table_structure.header_row_index,
            max_row=table_structure.header_row_index,
            values_only=True
        ))[0]
        
        # Convert headers to strings, use placeholder for empty headers
        headers = [
            str(h).strip() if h and str(h).strip() else f"Column_{i}" 
            for i, h in enumerate(header_row)
        ]
        
        logger.info(f"Extracted {len(headers)} headers")
        
        # Step 3: Match headers (THREE-TIER STRATEGY)
        logger.debug("Step 3: Matching headers using three-tier strategy")
        header_mappings = get_matcher().match_headers(headers)
        
        logger.info(f"Header matching complete: {len(header_mappings)} mappings created")
        
        # Step 4: Parse data rows
        logger.debug("Step 4: Parsing data rows")
        parsed_data: List[List[ParsedCell]] = []
        
        for row_idx in range(table_structure.data_start_row, sheet.max_row + 1):
            row_values = list(sheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                values_only=True
            ))[0]
            
            # Parse the row with header mappings
            parsed_cells = parse_row(row_values, header_mappings, row_index=row_idx)
            parsed_data.append(parsed_cells)
        
        logger.info(f"Parsed {len(parsed_data)} data rows")
        
        # Step 5: Compile results
        logger.debug("Step 5: Compiling results")
        total_cells = sum(len(row) for row in parsed_data)
        successful_parses = sum(
            1 for row in parsed_data for cell in row if cell.parse_success
        )
        
        # Count LLM calls (should be 0 or 1 per file)
        llm_matched_count = sum(1 for m in header_mappings if m.method.value == "llm")
        llm_calls_made = 1 if llm_matched_count > 0 else 0
        
        try:
            result = ParseResult(
                file_name=file.filename,
                table_structure=table_structure,
                header_mappings=header_mappings,
                parsed_data=parsed_data,
                total_cells=total_cells,
                successful_parses=successful_parses,
                llm_calls_made=llm_calls_made
            )
        except ValidationError as e:
            logger.error(f"Pydantic validation error creating ParseResult: {e}")
            raise HTTPException(
                status_code=422,
                detail=f"Data validation error: {str(e)}"
            )
        
        logger.info(f"Parsing complete: {total_cells} cells, "
                   f"{successful_parses} successful, {llm_calls_made} LLM calls")
        
        return result
        
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except ValidationError as e:
        logger.error(f"Pydantic validation error: {e}")
        raise HTTPException(
            status_code=422,
            detail=f"Data validation error: {str(e)}"
        )
    except openpyxl.utils.exceptions.InvalidFileException as e:
        logger.error(f"Invalid Excel file: {e}")
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid or corrupted Excel file: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, 
            detail=f"Error processing file: {str(e)}"
        )
