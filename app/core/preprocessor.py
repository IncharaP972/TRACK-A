"""
Excel preprocessing utilities for header row detection and table structure analysis.

This module provides functions to detect header rows in Excel files using
heuristics and LLM fallback, and to identify merged cells and table structure.
"""

import openpyxl
from typing import Optional, List, Tuple
from app.schema.models import TableStructure
from app.services.llm import LLMService
import logging

# Configure logging
logger = logging.getLogger(__name__)


def find_header_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet, 
    llm_service: Optional[LLMService] = None
) -> TableStructure:
    """
    Detect header row using heuristic or LLM fallback.
    
    Heuristic: Row with >3 string values is likely a header.
    If multiple candidates exist, selects the row with the most string cells.
    If ambiguous and LLM service is available, uses LLM for detection.
    
    Args:
        sheet: openpyxl worksheet to analyze
        llm_service: Optional LLM service for ambiguous cases
        
    Returns:
        TableStructure with detected header row, data start row, and merged cells
        
    Raises:
        ValueError: If sheet is invalid or has no data
    """
    logger.info("Starting header row detection")
    
    # Validate sheet
    if sheet is None:
        raise ValueError("Sheet cannot be None")
    
    if sheet.max_row < 1 or sheet.max_column < 1:
        raise ValueError(f"Sheet has no data: {sheet.max_row} rows, {sheet.max_column} columns")
    
    try:
        max_row_to_check = min(10, sheet.max_row)
        candidate_rows: List[Tuple[int, int, Tuple]] = []
        
        # Heuristic: Count string cells per row
        for row_idx in range(1, max_row_to_check + 1):
            try:
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                
                if string_count > 3:
                    candidate_rows.append((row_idx, string_count, row))
                    logger.debug(f"Row {row_idx} has {string_count} string cells - candidate header")
            except Exception as e:
                logger.warning(f"Error processing row {row_idx}: {e}")
                continue
        
        # Determine header row index
        if len(candidate_rows) == 1:
            # Clear winner
            header_row_idx = candidate_rows[0][0]
            logger.info(f"Single candidate found: row {header_row_idx}")
        elif len(candidate_rows) > 1:
            # Multiple candidates - use row with most strings
            header_row_idx = max(candidate_rows, key=lambda x: x[1])[0]
            logger.info(f"Multiple candidates found, selected row {header_row_idx} with most strings")
        else:
            # No clear candidates - use LLM if available
            if llm_service:
                logger.info("No clear candidates, using LLM for detection")
                try:
                    header_row_idx = _llm_detect_header_row(sheet, llm_service, max_row_to_check)
                except Exception as e:
                    logger.error(f"LLM header detection failed: {e}")
                    logger.warning("Defaulting to row 1")
                    header_row_idx = 1
            else:
                # Default to row 1
                logger.warning("No clear candidates and no LLM service, defaulting to row 1")
                header_row_idx = 1
        
        # Detect merged cells
        merged_cells = []
        try:
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col
                })
            logger.info(f"Detected {len(merged_cells)} merged cell ranges")
        except Exception as e:
            logger.warning(f"Error detecting merged cells: {e}")
            # Continue without merged cell information
        
        # Calculate data start row (header row + 1)
        data_start_row = header_row_idx + 1
        
        table_structure = TableStructure(
            header_row_index=header_row_idx,
            data_start_row=data_start_row,
            column_count=sheet.max_column,
            merged_cells=merged_cells
        )
        
        logger.info(f"Table structure detected: header at row {header_row_idx}, "
                    f"data starts at row {data_start_row}, {sheet.max_column} columns")
        
        return table_structure
        
    except ValueError:
        # Re-raise validation errors
        raise
    except Exception as e:
        logger.error(f"Critical error in find_header_row: {e}", exc_info=True)
        raise ValueError(f"Failed to detect header row: {str(e)}")


def _llm_detect_header_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet, 
    llm_service: LLMService, 
    max_rows: int
) -> int:
    """
    Use cheap LLM call to detect header row from first N rows.
    
    Args:
        sheet: openpyxl worksheet to analyze
        llm_service: LLM service for detection
        max_rows: Maximum number of rows to analyze
        
    Returns:
        Detected header row index (1-based)
        
    Raises:
        ValueError: If LLM detection fails completely
    """
    try:
        rows_text = []
        for row_idx in range(1, max_rows + 1):
            try:
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                rows_text.append(f"Row {row_idx}: {row}")
            except Exception as e:
                logger.warning(f"Error reading row {row_idx} for LLM detection: {e}")
                continue
        
        if not rows_text:
            raise ValueError("No rows could be read for LLM detection")
        
        prompt = f"""Given these Excel rows, which row number contains the column headers?

{chr(10).join(rows_text)}

Return only the row number as an integer."""
        
        logger.info("Querying LLM for header row detection")
        response = llm_service.simple_query(prompt)
        
        if not response:
            raise ValueError("LLM returned empty response")
        
        try:
            header_row_idx = int(response.strip())
            
            # Validate the row index is reasonable
            if header_row_idx < 1 or header_row_idx > max_rows:
                logger.warning(f"LLM returned out-of-range row index: {header_row_idx}, using 1")
                return 1
            
            logger.info(f"LLM detected header row: {header_row_idx}")
            return header_row_idx
        except ValueError as e:
            logger.error(f"Failed to parse LLM response as integer: {response}")
            raise ValueError(f"LLM response could not be parsed as integer: {response}")
            
    except Exception as e:
        logger.error(f"Error in LLM header detection: {e}")
        raise
